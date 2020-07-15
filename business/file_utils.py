'''
封装不同解析文件的方法
'''
import json
from openpyxl import load_workbook
# 解析json 文件 返回 python格式内容
def parse_from_json(filepath,tests_data):
    data = json.load(open(filepath, mode='r', encoding='utf8'))
    test_data = data[tests_data]
    return  test_data

# 解析 execl文件
def parse_from_excel(filepath,sheetname='Sheet1'):
    # 加载文件
    wb = load_workbook(filepath)
    # print(wb.worksheets)
    # 获得文档数据 test_data
    ws = wb[sheetname]
    test_data = []
    # 遍历
    for x in range(2, len(tuple(ws.rows)) + 1):
        test_case = []
        for y in range(2, 7):
            test_case.append(ws.cell(row=x, column=y).value)
        test_data.append(test_case)
    return test_data